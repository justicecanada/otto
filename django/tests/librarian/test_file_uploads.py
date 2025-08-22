import io
import os
from unittest.mock import Mock, patch

from django.core.files.uploadedfile import SimpleUploadedFile

import pytest
from docx import Document
from openpyxl import Workbook
from pptx import Presentation

from librarian.utils.process_engine import (
    csv_to_markdown,
    docx_to_markdown,
    excel_to_markdown,
    pdf_to_text_pymupdf,
    pptx_to_markdown,
)


def test_docx_to_markdown_success(sample_docx):
    result = docx_to_markdown(sample_docx)
    assert "Test Heading" in result
    assert "Test paragraph" in result


def test_docx_to_markdown_corrupted():
    with pytest.raises(Exception) as exc_info:
        docx_to_markdown(b"corrupted content")
    assert "Corrupt docx file" in str(exc_info.value)


def test_pptx_to_markdown_success(sample_pptx):
    result = pptx_to_markdown(sample_pptx)
    assert "Test Slide" in result
    assert "Test Content" in result


def test_pptx_to_markdown_corrupted():
    with pytest.raises(Exception) as exc_info:
        pptx_to_markdown(b"corrupted content")
    assert "Corrupt pptx file" in str(exc_info.value)


def test_excel_to_markdown_success(sample_excel):
    result = excel_to_markdown(sample_excel)
    assert "Sheet1" in result
    assert "Header1" in result
    assert "Header2" in result
    assert "Value1" in result
    assert "Value2" in result


def test_excel_to_markdown_corrupted():
    with pytest.raises(Exception) as exc_info:
        excel_to_markdown(b"corrupted content")
    assert "Corrupt Excel file" in str(exc_info.value)


def test_csv_to_markdown_success(sample_csv):
    result = csv_to_markdown(sample_csv)
    assert "Header1" in result
    assert "Header2" in result
    assert "Value1" in result
    assert "Value2" in result


def test_csv_to_markdown_corrupted():
    corrupted_content = b"\x80\x81\x82\x83"  # Invalid UTF-8 bytes
    with pytest.raises(Exception) as exc_info:
        csv_to_markdown(corrupted_content)
    assert "Corrupt CSV file" in str(exc_info.value)


def test_pdf_to_text_corrupted():
    with pytest.raises(Exception) as exc_info:
        pdf_to_text_pymupdf(b"corrupted content")
    assert "Corrupt PDF file" in str(exc_info.value)


def test_empty_csv():
    result = csv_to_markdown(b"")
    assert result == ""


def test_empty_excel():
    wb = Workbook()
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    result = excel_to_markdown(buffer.getvalue())
    assert "Sheet" in result  # Default sheet name


def test_markdown_table_formatting():
    csv_content = b"Header1,Header2\nValue1,Value2"
    result = csv_to_markdown(csv_content)
    expected = "| Header1 | Header2 |\n| --- | --- |\n| Value1 | Value2 |"
    assert result.strip() == expected.strip()


def test_multiple_excel_sheets(sample_excel):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1["A1"] = "Header1"
    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "Header2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    result = excel_to_markdown(buffer.getvalue())
    assert "# Sheet1" in result
    assert "# Sheet2" in result
    assert "Header1" in result
    assert "Header2" in result
